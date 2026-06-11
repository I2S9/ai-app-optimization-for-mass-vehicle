"""Serve styled single-sheet .xlsx exports built from pre-generated templates.

The templates (api/export_templates/*.xlsx) are produced offline by
`tools.build_export_templates` from the master workbook and already carry the
exact original styling with formula cells materialised to values. At runtime we
load a template once (cached in memory), overlay the user's current edits, save
to bytes, then restore the cached cells so the next request starts clean.
"""

from __future__ import annotations

import threading
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "export_templates"

_FILES = {
    "bd": "bd-export-template.xlsx",
    "synthesis": "synthesis-export-template.xlsx",
}
_DOWNLOAD_NAME = {
    "bd": "Database.xlsx",
    "synthesis": "Synthesis.xlsx",
}

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_cache: dict[str, openpyxl.Workbook] = {}
_locks: dict[str, threading.Lock] = {key: threading.Lock() for key in _FILES}


def _load_workbook(sheet: str) -> openpyxl.Workbook:
    """Parse the template once and keep the Workbook in module memory."""
    if sheet not in _cache:
        path = TEMPLATE_DIR / _FILES[sheet]
        if not path.is_file():
            raise FileNotFoundError(
                f"Export template missing: {path}. Build it with "
                "`cd api && python -m tools.build_export_templates`."
            )
        _cache[sheet] = openpyxl.load_workbook(path)
    return _cache[sheet]


def _coerce(value, number_format: str):
    """Turn an edit string into the right Python type so number formats render.

    Cells formatted as text ('@') keep their string verbatim — this protects
    zero-padded codes (e.g. "007") from being collapsed to an int.
    """
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value).strip()
    if s == "":
        return None
    if number_format and "@" in number_format:
        return s
    neg = s[1:] if s.startswith("-") else s
    if neg.isdigit():
        return int(s)
    try:
        return float(s)
    except ValueError:
        return s


def build_export_bytes(sheet: str, overrides: list[dict]) -> tuple[bytes, str]:
    """Apply `overrides` onto the cached template, return (xlsx_bytes, filename).

    overrides: [{"r": <1-based row int>, "c": "<column letter>", "v": <value>}, ...]

    The apply -> save -> restore section is serialised per sheet with a threading
    lock (the FastAPI route is sync, so it runs in Starlette's threadpool and
    concurrent calls are real threads). Writing `cell.value` leaves the cell's
    style index untouched, so all formatting is preserved.
    """
    wb = _load_workbook(sheet)
    ws = wb.worksheets[0]
    with _locks[sheet]:
        saved: list[tuple[int, int, object]] = []
        try:
            for ov in overrides:
                row = int(ov["r"])
                col = column_index_from_string(str(ov["c"]).upper())
                cell = ws.cell(row=row, column=col)
                saved.append((row, col, cell.value))
                cell.value = _coerce(ov.get("v"), cell.number_format)
            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue(), _DOWNLOAD_NAME[sheet]
        finally:
            for row, col, old in saved:
                ws.cell(row=row, column=col).value = old
