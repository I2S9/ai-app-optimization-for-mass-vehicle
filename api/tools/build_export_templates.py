"""Pre-build trimmed, value-materialized single-sheet Excel export templates.

These templates are what `app.excel_export` serves at runtime: one workbook
holding ONLY the BD sheet (and one holding ONLY SYNTHESIS), with every formula
already replaced by its cached value and ALL original styling preserved (fills,
fonts, borders, number formats, column widths, row heights, merged cells).

Run offline, one-time, and again whenever the master workbook changes. It needs
the 22 MB master `.xlsm`, which is gitignored AND .dockerignored — so this CANNOT
run inside `docker build`. Generate locally, then commit `api/export_templates/*.xlsx`
so the Dockerfile's `COPY api api` ships them.

    cd api && python -m tools.build_export_templates
"""

from __future__ import annotations

import logging
import re
import sys
import time
import zipfile
from pathlib import Path

import openpyxl

log = logging.getLogger("build_export_templates")

# api/tools/build_export_templates.py -> repo root is three parents up.
ROOT = Path(__file__).resolve().parent.parent.parent
WORKBOOK = ROOT / "workbooks" / "base-de-donnees-complete-avec-liens.xlsm"
OUT_DIR = Path(__file__).resolve().parent.parent / "export_templates"

# Excel tab name -> output filename. The tab names must match the workbook exactly.
TARGETS = {
    "BD": "bd-export-template.xlsx",
    "SYNTHESIS": "synthesis-export-template.xlsx",
}


def _used_dims(ws) -> tuple[int, int]:
    return ws.max_row, ws.max_column


def _cf_stats(ws) -> tuple[int, int]:
    """Return (rule_count, cross_sheet_operand_count) for a sheet's conditional
    formatting. openpyxl preserves *standard* CF rules (cellIs/colorScale/...) on
    save; only the x14 extension subset is dropped. A rule whose formula operand
    references another sheet ('!') would show #REF! once we delete that sheet."""
    rules = 0
    cross = 0
    try:
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                rules += 1
                for formula in (getattr(rule, "formula", None) or []):
                    if isinstance(formula, str) and "!" in formula:
                        cross += 1
    except Exception:  # noqa: BLE001 - openpyxl internals vary by version
        pass
    return rules, cross


def _strip_cruft(wb, ws) -> None:
    """Remove workbook-global structures left over from the deleted sheets that
    make Excel report "file is corrupt" (openpyxl/LibreOffice tolerate them).

    Stripped: all defined names (~12k, mostly #REF!), external workbook links
    (242 parts + their rels), charts/images/drawings, the legacy comments VML
    drawing, and per-cell comments. Everything VISIBLE on the sheet (fills, fonts,
    borders, number formats, column widths, row heights, merges, conditional
    formatting) is left untouched.
    """
    # Defined names — global and sheet-local.
    try:
        wb.defined_names.clear()
    except Exception:  # noqa: BLE001
        pass
    try:
        ws.defined_names.clear()
    except Exception:  # noqa: BLE001
        pass
    # External workbook links.
    try:
        wb._external_links = []
    except Exception:  # noqa: BLE001
        pass
    # Drawings / charts / images / legacy comments VML on the surviving sheet.
    ws._charts = []
    ws._images = []
    ws._drawing = None
    ws.legacy_drawing = None
    # Per-cell comments (their part + VML are a dangling-reference corruption source).
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None


def _verify_clean(path: Path, expected_cf: int, expected_merges: int) -> None:
    """Reopen the saved file and fail loudly if any Excel-corrupting cruft remains
    or if the visible structure (CF rules, merges) was not preserved."""
    z = zipfile.ZipFile(path)
    names = z.namelist()
    wbxml = z.read("xl/workbook.xml").decode("utf-8", "ignore")
    # A built-in name referencing only the surviving sheet (e.g. _xlnm._FilterDatabase
    # for the AutoFilter range) is valid and NOT a corruption source — Excel keeps
    # those. Flag only defined names that are #REF! or reference a deleted sheet.
    survivor_m = re.search(r'<sheet [^>]*name="([^"]+)"', wbxml)
    survivor = survivor_m.group(1) if survivor_m else None
    bad_defined = 0
    for body in re.findall(r"<definedName\b[^>]*>(.*?)</definedName>", wbxml, re.S):
        if "#REF!" in body:
            bad_defined += 1
            continue
        for ref_sheet in re.findall(r"'?([^,'!()+\-*/ ]+)'?!", body):
            if ref_sheet.strip() and ref_sheet.strip() != survivor:
                bad_defined += 1
                break
    offenders = {
        "externalLinks": sum(n.startswith("xl/externalLinks") for n in names),
        "drawings": sum(n.startswith("xl/drawings") for n in names),
        "charts": sum(n.startswith("xl/charts") for n in names),
        "vml": sum(n.endswith(".vml") for n in names),
        "comments": sum("comments" in n.lower() for n in names),
        "definedName(broken/cross-sheet)": bad_defined,
    }
    bad = {k: v for k, v in offenders.items() if v}
    if bad:
        raise SystemExit(
            f"{path.name}: leftover cruft after cleanup {bad} — Excel will reject it. "
            "Strengthen _strip_cruft (or fall back to a fresh-workbook rebuild)."
        )
    out_cf, _ = _cf_stats(openpyxl.load_workbook(path).worksheets[0])
    out_merges = len(openpyxl.load_workbook(path).worksheets[0].merged_cells.ranges)
    if out_cf != expected_cf:
        raise SystemExit(f"{path.name}: CF rules {out_cf} != expected {expected_cf}")
    if out_merges != expected_merges:
        raise SystemExit(
            f"{path.name}: merges {out_merges} != expected {expected_merges}"
        )
    log.info(
        "verified clean: %s (no externalLinks/definedNames/drawings/vml/comments; "
        "CF=%d, merges=%d)",
        path.name, out_cf, out_merges,
    )


def _audit(ws) -> None:
    """Log anything that won't survive into the export so it's never silent."""
    charts = getattr(ws, "_charts", None) or []
    images = getattr(ws, "_images", None) or []
    if charts or images:
        log.warning(
            "Sheet %r carries %d chart(s) / %d image(s) — openpyxl DROPS these on "
            "save. Confirm none are needed in the exported file.",
            ws.title, len(charts), len(images),
        )
    # Count formula cells that Excel never cached (None under data_only) — they
    # export blank. If this is high, recalc + re-save the master in Excel first.
    max_row, max_col = _used_dims(ws)
    blanks = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is None and cell.data_type == "f":
                blanks += 1
    if blanks:
        log.warning(
            "Sheet %r has %d formula cell(s) with no cached value (export blank). "
            "Recalculate and re-save the master workbook in Excel to populate them.",
            ws.title, blanks,
        )


def build_one(target: str, out_name: str) -> None:
    t0 = time.time()
    # data_only=True  -> formula cells come back as their cached computed value.
    # keep_vba=False  -> drop macros (none present) and produce a clean .xlsx.
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=False)
    log.info("loaded %s in %.1fs (%d sheets)", WORKBOOK.name, time.time() - t0, len(wb.sheetnames))

    if target not in wb.sheetnames:
        raise SystemExit(f"Sheet {target!r} not found. Tabs: {wb.sheetnames}")

    ws = wb[target]
    _audit(ws)
    src_cf, cross_cf = _cf_stats(ws)
    src_merges = len(ws.merged_cells.ranges)
    if cross_cf:
        log.warning(
            "Sheet %r has %d conditional-formatting operand(s) that reference "
            "another sheet — those rules may render #REF! once the other sheets "
            "are removed. Consider exporting the whole workbook for these.",
            ws.title, cross_cf,
        )

    # Keep only the target sheet. Deleting the others preserves the survivor's
    # styles/merges/dimensions, and cross-sheet formula references no longer
    # matter because data_only already materialised every value.
    for name in list(wb.sheetnames):
        if name != target:
            del wb[name]

    # Remove the leftover defined names / external links / drawings / comments
    # that make Excel reject the file as corrupt (see Addendum 2 in the plan).
    _strip_cruft(wb, ws)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / out_name
    wb.save(out_path)
    wb.close()

    # Hard gate: the saved file must be free of Excel-corrupting cruft and must
    # have kept its conditional formatting + merges. _verify_clean raises on fail.
    _verify_clean(out_path, expected_cf=src_cf, expected_merges=src_merges)

    log.info(
        "wrote %s (%s rows x %s cols, %.0f KB)",
        out_path.name, ws.max_row, ws.max_column, out_path.stat().st_size / 1024,
    )


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if not WORKBOOK.is_file():
        raise SystemExit(
            f"Master workbook missing: {WORKBOOK}\n"
            "Place it under workbooks/ (it is gitignored) and re-run."
        )
    for target, out_name in TARGETS.items():
        build_one(target, out_name)
    log.info("done -> %s", OUT_DIR)
    return 0


if __name__ == "__main__":
    sys.exit(main())
